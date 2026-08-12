from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BRAND = "e5006e"
DARK = "1a1a2e"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=DARK, size=14)
LABEL_FONT = Font(name="Calibri", bold=True, color=DARK, size=11)
VALUE_FONT = Font(name="Calibri", color="000000", size=11)
TOTAL_FONT = Font(name="Calibri", bold=True, color=BRAND, size=12)
HIGHLIGHT_FILL = PatternFill("solid", fgColor="fde8f2")
THIN_BORDER = Border(
    left=Side(style="thin", color="d0d0d8"),
    right=Side(style="thin", color="d0d0d8"),
    top=Side(style="thin", color="d0d0d8"),
    bottom=Side(style="thin", color="d0d0d8"),
)


def _money(v: float) -> float:
    return round(float(v or 0), 2)


def generate_payroll_xlsx(record, profile: dict) -> bytes:
    """record — PayrollRecord (ORM), profile — профиль пользователя из auth-service."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт ЗП"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24

    ws.merge_cells("A1:B1")
    ws["A1"] = "Бит.Serves — Расчёт заработной платы"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    info_rows = [
        ("Получатель", profile.get("full_name") or "—"),
        ("Отдел", profile.get("department_name") or "—"),
        ("Должность", profile.get("position_name") or "—"),
        ("Грейд", record.grade_name or record.grade_id),
        ("Период", record.period),
        ("Дата расчёта", record.created_at.strftime("%d.%m.%Y %H:%M") if record.created_at else "—"),
    ]
    row = 2
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2, value=str(value)).font = VALUE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Параметры расчёта").font = Font(bold=True, size=12, color=BRAND)
    row += 1

    is_abt = (getattr(record, "scheme", "margin") or "margin") == "abt"

    if is_abt:
        params = [
            ("Отработано дней", record.worked_days, "дн."),
            ("Рабочих дней в месяце", record.working_days, "дн."),
            ("Реализация: новые продажи", _money(record.sales_new), "₽"),
            ("Реализация: расширение", _money(record.sales_expansion), "₽"),
            ("Реализация: апгрейд", _money(record.sales_upgrade), "₽"),
            ("Реализация: продление без изменений", _money(record.sales_renew), "₽"),
            ("Товары СБИС", _money(record.sbis_goods), "₽"),
            ("НДФЛ", float(record.tax_rate), "%"),
        ]
        if record.has_plan and record.plan_margin:
            params.append(("Норма реализации ДС (план)", _money(record.plan_margin), "₽"))
            if record.performance_pct is not None:
                params.append(("Выполнение нормы", float(record.performance_pct), "%"))
    else:
        params = [
            ("Отработано дней", record.worked_days, "дн."),
            ("Рабочих дней в месяце", record.working_days, "дн."),
            ("Маржа за месяц (для плана и ступеней)", _money(record.month_margin), "₽"),
            ("Маржа с услуг", _money(record.service_margin), "₽"),
            ("Маржа с товара", _money(record.goods_margin), "₽"),
            ("Процент премии", float(record.bonus_percent), "%"),
            ("Коэффициент услуг", float(record.service_factor), ""),
            ("НДФЛ", float(record.tax_rate), "%"),
        ]

    if _money(record.kpi3_as_revenue) > 0:
        params.append(("KPI3 — приход с новых АС (без НДС)", _money(record.kpi3_as_revenue), "₽"))
        params.append(("KPI3 — процент премии", 5.0, "%"))
    if _money(record.kpi2_revenue) > 0:
        params.append(("KPI2 — приход ден. средств", _money(record.kpi2_revenue), "₽"))
        params.append(("KPI2 — сохранность клиентов", float(record.kpi2_retention_pct), "%"))
        if record.kpi2_enabled:
            params.append(("KPI2 — процент премии", float(record.grade_kpi2_bonus_percent), "%"))
            params.append(("KPI2 — мин. сохранность", float(record.grade_kpi2_min_retention_pct), "%"))
    elif is_abt and record.kpi2_enabled:
        params.append(("KPI2 — сохранность клиентов", float(record.kpi2_retention_pct), "%"))
        params.append(("KPI2 — мин. сохранность", float(record.grade_kpi2_min_retention_pct), "%"))
    for label, value, unit in params:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right")
        if unit == "₽":
            ws.cell(row=row, column=2).number_format = "#,##0.00"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Результат расчёта").font = Font(bold=True, size=12, color=BRAND)
    row += 1

    headers = ["Показатель", "Сумма, ₽"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    total_bonus_with_kpi2 = round(_money(record.bonus_total) + _money(record.kpi2_bonus_amount) + _money(record.kpi3_bonus_amount), 2)

    if is_abt:
        results = [
            ("Начислено по окладу", _money(record.accrued_base)),
            ("Премия: новые продажи", _money(record.bonus_new)),
            ("Премия: расширение", _money(record.bonus_expansion)),
            ("Премия: апгрейд", _money(record.bonus_upgrade)),
            ("Премия: продление без изменений", _money(record.bonus_renew)),
            ("Премия: товары СБИС (10%)", _money(record.bonus_sbis_goods)),
        ]
    else:
        results = [
            ("Начислено по окладу", _money(record.accrued_base)),
            ("Премия за услуги", _money(record.services_bonus)),
            ("Премия за товар", _money(record.goods_bonus)),
        ]

    if _money(record.kpi3_bonus_amount) > 0:
        results.append(("Премия за новые АС (KPI3, 5%)", _money(record.kpi3_bonus_amount)))
    if _money(record.kpi2_bonus_amount) > 0:
        results.append(("Премия за сохранность (KPI2)", _money(record.kpi2_bonus_amount)))
    elif record.kpi2_paid is False and (_money(record.kpi2_revenue) > 0 or (is_abt and float(record.kpi2_retention_pct or 0) > 0)):
        results.append(("Премия за сохранность (KPI2) — не выплач.", 0.0))

    results.extend([
        ("Премия итого", total_bonus_with_kpi2),
        ("Начислено всего (gross)", _money(record.gross_pay)),
        ("НДФЛ", _money(record.tax_amount)),
    ])
    for label, value in results:
        ws.cell(row=row, column=1, value=label).font = VALUE_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = "#,##0.00"
        row += 1

    ws.cell(row=row, column=1, value="К выплате (net)").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = HIGHLIGHT_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    cell = ws.cell(row=row, column=2, value=_money(record.net_pay))
    cell.font = TOTAL_FONT
    cell.fill = HIGHLIGHT_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = "#,##0.00"
    ws.row_dimensions[row].height = 24

    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
